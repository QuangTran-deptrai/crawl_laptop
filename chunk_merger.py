import pandas as pd
import os
import glob
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def clean_price(val):
    if pd.isna(val):
        return val
    val_str = str(val)
    digits = re.sub(r'[^\d]', '', val_str)
    if digits:
        return int(digits)
    return val

def merge_chunks():
    # Các trang web cần gộp
    stores = ["fptshop", "phongvu", "tgdd", "cellphones", "gearvn"]
    
    for store in stores:
        print(f"=== Đang gộp dữ liệu cho {store.upper()} ===")
        # Tìm tất cả file chunk của store này
        pattern = f"laptop_{store}_chunk_*.xlsx"
        chunk_files = glob.glob(pattern)
        
        if not chunk_files:
            print(f"Không tìm thấy file chunk nào cho {store}.")
            continue
            
        print(f"Tìm thấy {len(chunk_files)} mảnh: {chunk_files}")
        
        all_data = []
        for file in chunk_files:
            try:
                df = pd.read_excel(file)
                all_data.append(df)
            except Exception as e:
                print(f"Lỗi đọc file {file}: {e}")
                
        if all_data:
            # Nối tất cả dữ liệu
            merged_df = pd.concat(all_data, ignore_index=True)
            # Xóa dòng trùng lặp (nếu có, dựa vào URL/Link)
            # URL column có thể là 'URL' hoặc 'Link Sản Phẩm'
            if 'URL' in merged_df.columns:
                merged_df = merged_df.drop_duplicates(subset=['URL'])
            elif 'Link Sản Phẩm' in merged_df.columns:
                merged_df = merged_df.drop_duplicates(subset=['Link Sản Phẩm'])
                
            # Làm sạch giá tiền thành số
            price_cols = [c for c in merged_df.columns if 'giá' in c.lower() and 'giảm' not in c.lower() and 'khuyến' not in c.lower()]
            for col in price_cols:
                merged_df[col] = merged_df[col].apply(clean_price)
                
            output_file = f"laptop_{store}_all.xlsx"
            
            # Xuất và định dạng Excel
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                merged_df.to_excel(writer, index=False, sheet_name='Data')
                
                worksheet = writer.sheets['Data']
                
                # Định dạng Header
                header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                header_font = Font(bold=True)
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                # Căn lề, tự động xuống dòng và định dạng cột
                for col_idx, col_name in enumerate(merged_df.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    
                    # Chỉnh chiều rộng cột
                    if col_name.lower() in ['tên sản phẩm', 'cấu hình chi tiết', 'khuyến mãi']:
                        worksheet.column_dimensions[col_letter].width = 45
                    elif 'url' in col_name.lower() or 'link' in col_name.lower():
                        worksheet.column_dimensions[col_letter].width = 40
                    else:
                        worksheet.column_dimensions[col_letter].width = 18
                        
                    # Định dạng số tiền
                    if col_name in price_cols:
                        for row_idx in range(2, len(merged_df) + 2):
                            cell = worksheet[f"{col_letter}{row_idx}"]
                            if isinstance(cell.value, int):
                                cell.number_format = '#,##0"đ"'
                                
                # Tự động xuống dòng cho toàn bộ bảng
                for row in worksheet.iter_rows(min_row=2, max_col=len(merged_df.columns)):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        
            print(f"✅ Đã gộp và định dạng thành công {len(merged_df)} laptop vào {output_file}")
            
            # Xóa các file chunk rác
            for file in chunk_files:
                os.remove(file)
            print(f"Đã dọn dẹp {len(chunk_files)} file chunk rác.")
        print("-" * 40)

if __name__ == "__main__":
    merge_chunks()
